import numpy as np
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import pandas as pd
import streamlit as st
from datetime import datetime
import re
from dataclasses import dataclass, field
import openpyxl
from typing import Optional
import json
from openpyxl.reader.excel import load_workbook


class ExcelParsingErrors:
    def __init__(self):
        self.warnings = []
        self.errors = []
    warnings: list[str]
    errors: list[str]

    def warning(self, message: str):
        self.warnings.append(message)

    def error(self, message: str):
        self.errors.append(message)

    def areErrors(self):
        return len(self.errors) > 0






def parseQuantity(
    qty: str, row: int, buyer: str, errors: ExcelParsingErrors
) -> tuple[float, ExcelParsingErrors]:
    cleaned = re.sub(r"[^0-9\.]", "", qty)
    floatQty = 0.0
    try:
        floatQty = float(cleaned)
    except ValueError:
        errors.error(
            f"Quantity in row {row} for buyer {buyer} could not be parsed.")

    return floatQty, errors


def parseContactHeaders(
    headers: list, errors: ExcelParsingErrors
) -> tuple[dict[str, int], dict[int, str], ExcelParsingErrors]:
    column_mapping = {
        "Buyer Key as in Spreadsheet": "key",
        "Buyer Full Name": "name",
        "Address Line 1": "address1",
        "Address Line 2": "address2",
        "City": "city",
        "Postcode": "postcode",
        "Country": "country",
    }

    contact_header_dict = {
        "key": None,
        "name": None,
        "address1": None,
        "address2": None,
        "city": None,
        "postcode": None,
        "country": None,
    }
    buyers: dict[int, str] = {}

    for key, value in column_mapping.items():
        try:
            index_of_header = headers.index(key)

        except ValueError:
            errors.error(
                f"Header {key} could not be found in the contacts sheet.")

        except Exception:
            errors.error(f"Unknown error when parsing headers")

        contact_header_dict[value] = index_of_header

    return contact_header_dict, buyers, errors


def contacts_uploader(
    contacts_sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[list[Buyer], ExcelParsingErrors]:
    """
    This function takes in the contacts dataframe and returns a
    list of all the buyers.
    """

    parsing_errors = ExcelParsingErrors()

    header_row = 1

    headers = contacts_sheet[header_row]
    headers = [cell.value for cell in headers]
    headers_dict, buyers, parsing_errors = parseContactHeaders(
        headers, parsing_errors)

    buyers: list[Buyer] = []

    i = header_row - 1

    for row in contacts_sheet.iter_rows(min_row=header_row + 1, values_only=True):
        i = i + 1

        if row[headers_dict["key"]] == None or row[headers_dict["name"]] == "":
            continue

        country = row[headers_dict["country"]]
        if country == None:
            country = ""

        buyer = Buyer(
            key=row[headers_dict["key"]],
            name=row[headers_dict["name"]],
            address1=row[headers_dict["address1"]],
            city=row[headers_dict["city"]],
            postcode=row[headers_dict["postcode"]],
            country=country,
            address2=row[headers_dict["address2"]],
        )

        buyers.append(buyer)

    return buyers, parsing_errors


def orderify(
    order_sheet: openpyxl.worksheet.worksheet.Worksheet,
    buyers: list[Buyer],
    parsing_errors: ExcelParsingErrors,
    order_sheet_name=None,
) -> tuple[list[OrderLine], ExcelParsingErrors]:
    """
    This function takes in the marketplace dataframe and returns a
    dataframe with each order as a separate row.
    """

    header_row = 3

    headers = order_sheet[header_row]
    headers = [cell.value for cell in headers]
    headers_dict, buyers, parsing_errors = parseOrderHeaders(
        headers, buyers, parsing_errors)

    order_lines: list[OrderLine] = []

    i = header_row - 1
    for row in order_sheet.iter_rows(min_row=header_row+1, values_only=True):
        i = i + 1

        if (i < 10):
            print(row)

        if row[headers_dict["price"]] == None or row[headers_dict["price"]] == "":
            continue

        for index, buyer in buyers.items():
            qty, parsing_errors = parseQuantity(
                str(row[index]), i, buyer, parsing_errors
            )
            if qty == 0:
                continue

            order_line = OrderLine(
                produce=row[headers_dict["produce"]],
                variant=row[headers_dict["variant"]],
                unit=row[headers_dict["unit"]],
                seller=row[headers_dict["seller"]],
                qty=qty,
                buyer=buyer,
            )
            parsing_errors = order_line.setPrice(
                str(row[headers_dict["price"]]), i, parsing_errors
            )
            parsing_errors = order_line.checkErrors(i, parsing_errors)

            order_lines.append(order_line)

    # print(order_lines)
    print(parsing_errors.errors)

    return order_lines, parsing_errors


def contacts_formatter(contacts):
    contacts = contacts.replace({np.nan: ""})

    column_mapping = {
        "Buyer Key as in Spreadsheet": "key",
        "Buyer Full Name": "name",
        "Address Line 1": "address1",
        "Address Line 2": "address2",
        "City": "city",
        "Postcode": "postcode",
        "Country": "country",
    }

    keys = list(column_mapping.keys())

    # Check if all keys exist in the columns of orders
    if not set(keys).issubset(set(contacts.columns)):
        missing_columns = list(set(keys) - set(contacts.columns))
        missing_columns_str = ", ".join(
            missing_columns
        )  # Join the missing columns into a string
        print(
            f"Error: Missing columns in contacts spreadsheet: {missing_columns_str}")
        st.error(
            f"Error: Missing columns in contacts spreadsheet: {missing_columns_str}"
        )
        st.stop()

    contacts = contacts.rename(columns=column_mapping)
    # Add a new column called "number" with empty values
    contacts["number"] = ""
    st.toast(":white_check_mark: Contacts column names are correct")
    return contacts


def contacts_checker(contacts, buyers):
    unmatched_buyers = []
    for buyer in buyers:
        if buyer not in contacts.tolist():
            unmatched_buyers.append(buyer)
    if unmatched_buyers:
        st.error(
            f"Buyers: {unmatched_buyers} not found in contacts spreadsheet")
        print(f"Buyers: {unmatched_buyers} not found in contacts spreadsheet")
        st.stop()
    else:
        st.toast(":white_check_mark: All buyers found in contacts")
        print("All buyers found in contacts")
    return unmatched_buyers


def extract_buyer_list(orders):
    if "BUYERS:" not in orders.columns:
        st.error("Error: Missing column 'BUYERS:' in order spreadsheet")
        print("Error: Missing column 'BUYERS:' in order spreadsheet")
        st.stop()
    buyers_column_index = orders.columns.get_loc("BUYERS:")
    buyers = orders.columns[buyers_column_index + 1:][
        orders.iloc[:, buyers_column_index + 1:].sum() > 0
    ].tolist()
    print(f"Buyer list: {buyers}")
    return buyers


def date_extractor(order_sheet):
    try:
        order_sheet_name = str(order_sheet)
        date_str = order_sheet_name.split(" - ")[-1].split(".")[0]
        dateobj = datetime.strptime(date_str, "%d_%m_%Y")
        return dateobj
    except ValueError as e:
        st.error(
            f"Error: Failed to extract date from order sheet: {order_sheet.name}, make sure the file name is in the format '...N - dd_mm_yyyy.xlsx'"
        )
        print(
            f"Error: Failed to extract date from order sheet: {order_sheet.name}, make sure the file name is in the format '...N - dd_mm_yyyy.xlsx"
        )
        st.stop()




if __name__ == "__main__":
    spreadsheet_all = openpyxl.load_workbook("./test.xlsx", data_only=True)
    order_sheet = spreadsheet_all["GROWERS' PAGE"]

    orderify(order_sheet)
